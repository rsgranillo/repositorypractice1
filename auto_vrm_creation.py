#imports and modules
import openpyxl
import requests
import base64
import json
import jira
from requests.auth import HTTPBasicAuth
from datetime import date
from openpyxl import load_workbook
from atlassian import Jira
from jira import JIRA

#Requested start and end date
today = date.today()

#Get data from excel sheet
wb = load_workbook(filename = 'Auto VRM creation.xlsx')
ws = wb.get_sheet_by_name('Sheet1')

def iter_rows(ws):
    for row in ws.iter_rows():
        yield [cell.value for cell in row]

cells = list(iter_rows(ws))
#print(cells)

iterator_row = [1,2]
values_dict = {'software':[],
              'vehicle':[],
              'test_type':[],
              'test_plan':[]}
for i in range(len(iterator_row)):
    values_dict['software'].append(cells[i][1])
    values_dict['vehicle'].append(cells[i][0])
    values_dict['test_type'].append(cells[i][2])
    values_dict['test_plan'].append(cells[i][3])     
    i += 1
    
print(values_dict)

#create issue (s)

for i in range(len(iterator_row)):
    sw = values_dict['software'][i] #desired software
    vehicle = values_dict['vehicle'][i] #desired vehicle name
    req_type = values_dict['test_type'][i] #test driver request type
    test_plan = values_dict['test_plan'][i] #jama software test plan
    print("Logging into Jira..")
    url = "https://rivianautomotivellc.atlassian.net/rest/api/2/issue/"
    auth = HTTPBasicAuth("rubygranillo@rivian.com",
                         "Ly5PbJ0UuJWUvYlW15BIB5FD")

    get_details_url = url

    payload = '''{
        "fields": {
           "project":
           { 
              "key": "VRM"
           },
           "summary": "Not a real ticket",
           "issuetype": {
              "name": "Test Driver"
           }             
       }
    } ''' 

    headers = {
       "Accept": "application/json",
       "Content-Type": "application/json"
    }
    print("Creating JIRA ticket...")
    response = requests.request(
       "POST",
       get_details_url,
       data=payload,
       headers=headers,
       auth=auth
    )

    ticket = json.dumps(json.loads(response.text), sort_keys=True, indent=4, separators=(",", ": "))
    ticket_dict = json.loads(ticket)
    ticket_id = ticket_dict['key']
    print("Updating ticket fields...")

    issue = jira.issue(f'{ticket_id}')
    issue.update(fields={'summary': 'Not a real ticket', 
                         'description': 'Not a real ticket', 
                         'customfield_14179': { 'value': 'Continuous Driver Validation'}, 
                         'customfield_11878': { 'value': 'Software-Controls-Propulsion'},
                         'customfield_17183': { 'value': 'WaterWorks'},
                         'customfield_17243': { 'value': 'Yes'},
                         'customfield_11437': sw,
                         'customfield_12530': {'value': vehicle},
                         'customfield_14179': {'value': req_type},
                         'customfield_17935': {'value': test_plan}
                        })
    print(f"https://rivianautomotivellc.atlassian.net/browse/{ticket_id}")
    i += 1
#'customfield_11170': { 'value': f'{today}'},
#'customfield_11143': { 'value': f'{today}'},